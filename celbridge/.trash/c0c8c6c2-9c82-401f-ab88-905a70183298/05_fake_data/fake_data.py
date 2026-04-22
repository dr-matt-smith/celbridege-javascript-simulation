import sys
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Configuration
DEFAULT_COUNT = 20
FILENAME = "fake_data.xlsx"

# Parse optional argument for row count
if len(sys.argv) > 1:
    try:
        COUNT = int(sys.argv[1])
    except ValueError:
        print("Invalid argument. Please provide a number for the row count.")
        sys.exit(1)
else:
    COUNT = DEFAULT_COUNT

# Write the Excel file to the same directory as this script
outfile = Path(__file__).parent / FILENAME

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "People"

# Add header row
headers = ["Full Name", "Email", "Phone", "Address"]
ws.append(headers)

# Adjust column widths
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 36

# Apply header style
hdr_font = Font(bold=True)
hdr_fill = PatternFill("solid", fgColor="DDDDDD")
hdr_align = Alignment(horizontal="center")

for cell in ws[1]:
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align

# Generate fake data rows
fake = Faker()
for _ in range(COUNT):
    name = fake.name()
    email = fake.email()
    phone = fake.phone_number()
    addr = fake.street_address()
    ws.append([name, email, phone, addr])

# Save the file
wb.save(outfile)
print(f"Wrote {COUNT} fake people to {outfile}")
