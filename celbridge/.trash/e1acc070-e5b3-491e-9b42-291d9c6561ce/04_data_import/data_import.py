import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# Dublin Bikes GBFS API endpoints
GBFS_STATION_INFO_URL = "https://api.cyclocity.fr/contracts/dublin/gbfs/v2/station_information.json"
GBFS_STATION_STATUS_URL = "https://api.cyclocity.fr/contracts/dublin/gbfs/v2/station_status.json"
OUTPUT_XLSX = "04_data_import/dublinbikes.xlsx"

def main():

    # Download station information (names, locations)
    info_response = requests.get(GBFS_STATION_INFO_URL, timeout=15)
    info_response.raise_for_status()
    
    # Build a dictionary of stations by their ID
    station_info = {}
    for station in info_response.json()["data"]["stations"]:
        station_id = station["station_id"]
        station_info[station_id] = station

    # Download station status (real-time availability)
    status_response = requests.get(GBFS_STATION_STATUS_URL, timeout=15)
    status_response.raise_for_status()
    
    # Build a dictionary of station status by ID
    station_status = {}
    for station in status_response.json()["data"]["stations"]:
        station_id = station["station_id"]
        station_status[station_id] = station

    # Create workbook & sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Bikes"

    # Write the header row
    headers = ["Station Name", "Available Bikes", "Empty Slots"]
    ws.append(headers)

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15

    # Apply header style (bold, shaded, centered)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Combine station info and status
    stations = []
    for station_id, info in station_info.items():
        if station_id in station_status:
            status = station_status[station_id]
            stations.append({
                "name": info.get("name", ""),
                "bikes": status.get("num_bikes_available", 0),
                "docks": status.get("num_docks_available", 0)
            })
    
    # Sort stations alphabetically by name (case-insensitive)
    stations.sort(key=lambda s: s["name"].lower())

    # Write a row for each station
    for s in stations:
        ws.append([s["name"], s["bikes"], s["docks"]])

    # Save the Excel file
    wb.save(OUTPUT_XLSX)
    print(f"Saved {OUTPUT_XLSX} with {len(stations)} stations.")

if __name__ == "__main__":
    main()
